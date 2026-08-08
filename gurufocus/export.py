"""
שלב 9 — ייצוא.
================================================================================
אחריות יחידה: לכתוב את מה שכבר חושב. אין כאן שינוי נתונים מלבד סדר עמודות.

קובץ האקסל מכיל את כל מה שצריך כדי לשפוט את הריצה בלי להסתכל בקוד:

    Data       הפאנל עצמו
    Coverage   מיפוי השדות — מה נמצא, איפה, ואיך
    Nulls      אחוז ריקים לכל עמודה + הערך האחרון
    Checks     בדיקות התקינות
    Manifest   סיכום ריצה לכל טיקר
"""

from __future__ import annotations

import logging
from collections.abc import Collection, Sequence
from datetime import datetime
from pathlib import Path

import pandas as pd

from .decomposition import (
    DECOMPOSITION_DUMMY_COLUMNS,
    HIGH_PRECISION_DECOMPOSITION_COLUMNS,
    NOPAT_BRIDGE_COLUMNS,
    ROIC_BRIDGE_COLUMNS,
    TEXT_DECOMPOSITION_COLUMNS,
)
from .wacc import (
    HIGH_PRECISION_WACC_COLUMNS,
    TEXT_WACC_COLUMNS,
    WACC_CORE_COLUMNS,
)

log = logging.getLogger(__name__)


EXCLUDED_OUTPUT_COLUMNS = {
    "currency",
    "exchange",
    "reporting_unit",
    "form_type",
    "accession_number",
    "cik",
    "filing_url",
    "filing_match",
    "calendar_year",
    "calendar_quarter",
}

STRUCTURAL_OUTPUT_COLUMNS = [
    "symbol",
    "company",
    "sector",
    "industry",
    "fiscal_period_end_date",
    "filing_date",
    "period_key",
    "period_year",
    "period_quarter",
    "run_date",
]

# עמודות טקסט ובוליאניות: אסור להמיר אותן ל-float, אחרת כל סיווג הופך ל-NaN.
NON_NUMERIC_OUTPUT_COLUMNS = frozenset(TEXT_DECOMPOSITION_COLUMNS | TEXT_WACC_COLUMNS)

# תרומות, משקלים ושיעורי ריבית הם יחסים קטנים — בפורמט "0.00" תרומה של 31
# נקודות בסיס ועלות חוב של 2.7% היו מוצגות כאפס.
HIGH_PRECISION_OUTPUT_COLUMNS = frozenset(
    HIGH_PRECISION_DECOMPOSITION_COLUMNS | HIGH_PRECISION_WACC_COLUMNS
)
HIGH_PRECISION_NUMBER_FORMAT = "0.000000"

# הסדר שנקבע למחקר: מתחיל מיד אחרי run_date וללא כפילויות.
PRIMARY_OUTPUT_COLUMNS = [
    "tax_provision",
    "calc_tax_expense_quarterly",
    "pretax_income",
    "calc_raw_tax_rate_quarterly",
    "ebit",
    "calc_nopat_quarterly",
    # פירוק השינוי ב-NOPAT — צמוד ל-NOPAT עצמו.
    *NOPAT_BRIDGE_COLUMNS,
    "total_current_assets",
    "cash_and_cash_equivalents",
    "short_term_investments",
    "total_current_liabilities",
    "short_term_debt",
    "intangible_assets",
    "goodwill",
    "net_ppe",
    "calc_ic_raw",
    "calc_average_ic_raw_quarterly",
    # כל ROIC רבעוני מלווה בתאום השנתי שלו, (1+r)^4-1, כדי שההשוואה
    # מול WACC תהיה בין שני שיעורים על אותו אופק זמן.
    "calc_roic_pretax_quarterly_ic_raw",
    "calc_roic_pretax_annualized_ic_raw",
    "calc_roic_posttax_quarterly_ic_raw",
    "calc_roic_posttax_annualized_ic_raw",
    # פירוק השינוי ב-ROIC אחרי מס — צמוד ל-ROIC עצמו.
    *ROIC_BRIDGE_COLUMNS,
    # Debt-to-equity: every direct input precedes the calculation.
    "short_term_debt_and_capital_lease",
    "long_term_debt_and_capital_lease",
    "calc_debt_value_quarterly",
    "total_assets",
    "total_liabilities",
    "equity",
    "total_stockholders_equity",
    "calc_debt_to_equity_quarterly",
    "valuations__ratios__debt_to_equity",
    # EV/FCF: repeated inputs are intentional for auditability.
    "market_cap",
    "calc_debt_value_quarterly",
    "cash_and_cash_equivalents",
    "short_term_investments",
    "calc_enterprise_value_quarterly",
    "free_cash_flow",
    "calc_free_cash_flow_ttm",
    "calc_ev_to_fcf_quarterly",
    "valuations__valuation_ratios__enterprise_value_to_fcf",
    "valuations__per_share_data__shares_outstanding",
    "valuations__per_share_data__month_end_stock_price",
    # WACC: אחרי כל העמודות הקיימות. מקורות משותפים חוזרים בכוונה כדי
    # שאפשר יהיה לבקר את החישוב בלי לדלג בין קצות הגיליון.
    "market_cap",
    "calc_wacc_equity_value",
    "calc_debt_value_quarterly",
    "calc_wacc_average_debt",
    "calc_wacc_total_capital",
    "calc_wacc_equity_weight",
    "calc_wacc_debt_weight",
    "calc_wacc_risk_free_rate",
    "calc_wacc_equity_risk_premium",
    "calc_wacc_cost_of_equity",
    "interest_expense",
    "calc_interest_expense_ttm",
    "calc_wacc_cost_of_debt",
    "calc_wacc_tax_rate",
    "calc_wacc_after_tax_cost_of_debt",
    "calc_wacc_annual",
    "calc_wacc_quarterly",
    "calc_wacc_quality_flag",
    "calc_wacc_inputs_complete",
    "calc_roic_minus_wacc_annualized",
    "calc_roic_minus_wacc_quarterly",
    "calc_creates_value",
    # עמודות הדמה נשארות בסוף: הן מיועדות לצריכה תוכנתית ולא לקריאה.
    # ב-Excel הן עוברות לגיליון Decomposition נפרד.
    *DECOMPOSITION_DUMMY_COLUMNS,
]


# בלוק ה-WACC מסודר ידנית כדי שכל מקור יופיע צמוד לחישוב שצורך אותו, ולכן
# הוא לא נפרש אוטומטית כמו עמודות הפירוק. הבדיקה הזאת מוודאת שעמודה חדשה
# במודול ה-WACC לא תישאר מחוץ לגיליון בשקט.
_unplaced = [
    column for column in WACC_CORE_COLUMNS if column not in PRIMARY_OUTPUT_COLUMNS
]
if _unplaced:
    raise RuntimeError(
        "עמודות WACC שאינן מופיעות בסדר הפלט: " + ", ".join(_unplaced)
    )


def _research_order(
    frame: pd.DataFrame,
    *,
    repeat_sources: bool = False,
) -> pd.DataFrame:
    preferred = STRUCTURAL_OUTPUT_COLUMNS + PRIMARY_OUTPUT_COLUMNS
    candidates = preferred if repeat_sources else dict.fromkeys(preferred)
    ordered = [
        column
        for column in candidates
        if column in frame.columns and column not in EXCLUDED_OUTPUT_COLUMNS
    ]
    ordered_set = set(ordered)
    remaining = [
        column
        for column in frame.columns
        if column not in ordered_set and column not in EXCLUDED_OUTPUT_COLUMNS
    ]
    result = frame.loc[:, ordered + remaining].copy()
    # כל נתוני המחקר המספריים נשמרים כ-float. שדות המבנה הם מזהים,
    # תאריכים ותוויות תקופה ולכן נשמרים בסוג הטבעי שלהם.
    #
    # עמודות הסיווג של הפירוק הן טקסט ובוליאני. בלי החרגה מפורשת ההמרה
    # ל-float הייתה הופכת כל תווית כמו ROIC_INCREASE_MIXED_EFFECTS ל-NaN
    # בשקט, והפלט היה נראה תקין אך ריק.
    for position, column in enumerate(result.columns):
        if column in STRUCTURAL_OUTPUT_COLUMNS:
            continue
        if column in NON_NUMERIC_OUTPUT_COLUMNS:
            continue
        numeric = pd.to_numeric(result.iloc[:, position], errors="coerce")
        result.isetitem(position, numeric.astype(float))
    return result


def order_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """מסדר את העמודות: מפתחות -> יישור -> שדות מדווחים -> מחושבים.

    עמודה שקיימת בטבלה אך לא ברשימה מצורפת בסוף, כדי שלעולם לא נאבד נתון.
    """
    return _research_order(frame)


def calculation_audit_view(frame: pd.DataFrame) -> pd.DataFrame:
    """מחזיר תצוגת מחקר ובה כל מקורות החישוב צמודים לתוצאה.

    שדות מקור משותפים חוזרים בכוונה לפני חישובים שונים כדי שהפלט יהיה
    ניתן לביקורת גם ללא דילוג בין עמודות.
    """
    if frame.empty:
        return frame.copy()
    if not frame.columns.is_unique:
        raise ValueError("תצוגת המחקר מצפה לטבלה עם שמות עמודות ייחודיים")
    return _research_order(frame, repeat_sources=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def build_output_path(
    directory: Path, period: str, extension: str, *, suffix: str = ""
) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    stem = f"GuruFocus_{period}_{_timestamp()}{suffix}"
    path = directory / f"{stem}.{extension}"
    if not path.exists():
        return path

    # לא דורסים קובץ מאותה ריצה/יום — במיוחד כשהוא פתוח ב-Excel.
    version = 2
    while True:
        candidate = directory / f"{stem}_v{version}.{extension}"
        if not candidate.exists():
            return candidate
        version += 1


def split_decomposition_dummies(
    frame: pd.DataFrame,
    *,
    keys: Sequence[str] = ("symbol", "period_key"),
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """מפריד את עמודות הדמה מגיליון Data אל טבלה משלהן.

    שישים ושלוש עמודות של אפסים ואחדות אינן קריאות לאדם, אבל הן נחוצות
    לצריכה תוכנתית. לכן הן נשארות ב-CSV וב-parquet ועוברות ב-Excel לגיליון
    נפרד, מזוהות לפי טיקר ותקופה.
    """
    present = [
        column for column in DECOMPOSITION_DUMMY_COLUMNS if column in frame.columns
    ]
    if not present:
        return frame, pd.DataFrame()
    identifiers = [column for column in keys if column in frame.columns]
    dummies = frame.loc[:, identifiers + present].copy()
    data = frame.drop(columns=present)
    return data, dummies


def write_excel(
    path: Path,
    data: pd.DataFrame,
    *,
    coverage: pd.DataFrame | None = None,
    nulls: pd.DataFrame | None = None,
    checks: pd.DataFrame | None = None,
    manifest: pd.DataFrame | None = None,
    decomposition: pd.DataFrame | None = None,
) -> Path:
    """כותב את חוברת האקסל המלאה."""
    path.parent.mkdir(parents=True, exist_ok=True)
    sheets: list[tuple[str, pd.DataFrame]] = [("Data", data)]

    for name, table in (
        ("Decomposition", decomposition),
        ("Coverage", coverage),
        ("Nulls", nulls),
        ("Checks", checks),
        ("Manifest", manifest),
    ):
        if table is not None and not table.empty:
            sheets.append((name, table))

    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        for name, table in sheets:
            table.to_excel(writer, sheet_name=name, index=False)
            _autosize(writer.sheets[name], table)
            if name == "Data":
                _format_float_values(writer.sheets[name], table)

    log.info("נשמר: %s (%d שורות, %d עמודות)", path, len(data), data.shape[1])
    return path


def _autosize(worksheet, frame: pd.DataFrame, *, max_width: int = 42) -> None:
    """רוחב עמודות סביר — כדי שהקובץ יהיה קריא בלי התאמה ידנית."""
    from openpyxl.utils import get_column_letter

    for position, column in enumerate(frame.columns, start=1):
        header = len(str(column))
        # בחירת עמודה לפי מיקום חיונית: בתצוגת הביקורת יש בכוונה
        # כותרות כפולות כאשר אותו מקור משתתף בכמה חישובים.
        sample = frame.iloc[:200, position - 1].astype(str).str.len().max()
        width = max(header, int(sample) if pd.notna(sample) else 0) + 2
        worksheet.column_dimensions[get_column_letter(position)].width = min(width, max_width)


def number_format_for(column: str) -> str | None:
    """הפורמט המספרי של עמודה בגיליון Data, או None אם אין להחיל פורמט."""
    if column in STRUCTURAL_OUTPUT_COLUMNS or column in NON_NUMERIC_OUTPUT_COLUMNS:
        return None
    if column in HIGH_PRECISION_OUTPUT_COLUMNS:
        return HIGH_PRECISION_NUMBER_FORMAT
    return "0.00"


def _format_float_values(worksheet, frame: pd.DataFrame) -> None:
    """מציג את כל נתוני המחקר המספריים עם מספר הספרות המתאים להם."""
    for position, column in enumerate(frame.columns, start=1):
        number_format = number_format_for(column)
        if number_format is None:
            continue
        for row in range(2, len(frame) + 2):
            worksheet.cell(row=row, column=position).number_format = number_format


def write_csv(
    path: Path,
    data: pd.DataFrame,
    *,
    full_precision_columns: Collection[str] = HIGH_PRECISION_OUTPUT_COLUMNS,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    # float_format חל על כל הקובץ, ולכן תרומת ROIC של 0.0031 הייתה נכתבת
    # כ-0.00 ומאבדת את כל המידע. העמודות האלה מומרות למחרוזת בדיוק מלא
    # לפני הכתיבה, וכל השאר נשאר בשתי ספרות כמו קודם.
    #
    # הגישה היא לפי מיקום ולא לפי שם: בתצוגת הביקורת יש כותרות כפולות
    # במכוון, ושם עמודה כפול היה מחזיר DataFrame במקום Series.
    wanted = set(full_precision_columns)
    positions = [
        position
        for position, column in enumerate(data.columns)
        if column in wanted
    ]
    if positions:
        data = data.copy()
        for position in positions:
            values = pd.to_numeric(data.iloc[:, position], errors="coerce")
            data.isetitem(
                position,
                values.map(
                    lambda value: "" if pd.isna(value) else format(value, ".12g")
                ),
            )

    # utf-8-sig כדי שאקסל בעברית יפתח את הקובץ בקידוד הנכון
    data.to_csv(
        path,
        index=False,
        encoding="utf-8-sig",
        float_format="%.2f",
    )
    log.info("נשמר: %s", path)
    return path


def write_parquet(path: Path, data: pd.DataFrame) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        data.to_parquet(path, index=False)
    except ImportError as exc:
        raise RuntimeError(
            "כתיבת parquet דורשת pyarrow — התקינו: pip install pyarrow"
        ) from exc
    log.info("נשמר: %s", path)
    return path
