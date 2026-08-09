"""Tests for the reduced Data-sheet schema and display order."""

from __future__ import annotations

import csv

import pandas as pd
from openpyxl import load_workbook
from pandas.api.types import is_float_dtype

from gurufocus.calculations import add_calculated
from gurufocus.decomposition import DECOMPOSITION_DUMMY_COLUMNS
from gurufocus.export import (
    EXCLUDED_OUTPUT_COLUMNS,
    NON_NUMERIC_OUTPUT_COLUMNS,
    PRIMARY_OUTPUT_COLUMNS,
    STRUCTURAL_OUTPUT_COLUMNS,
    calculation_audit_view,
    number_format_for,
    order_columns,
    split_decomposition_dummies,
    write_csv,
    write_excel,
)
from gurufocus.extract import build_frame


def _exported_panel(quarterly_payload):
    frame, _ = build_frame(quarterly_payload, "TEST")
    frame["sector"] = "Technology"
    frame["industry"] = "Software"
    frame["market_cap"] = 1000.0
    frame["valuations__valuation_ratios__enterprise_value_to_fcf"] = 12.5
    frame["valuations__per_share_data__shares_outstanding"] = 50.0
    frame["valuations__per_share_data__month_end_stock_price"] = 20.0
    frame["valuations__ratios__debt_to_equity"] = 0.25
    frame["run_date"] = "2026-08-01"
    return calculation_audit_view(order_columns(add_calculated(frame)))


def test_data_columns_have_the_exact_requested_order(quarterly_payload):
    panel = _exported_panel(quarterly_payload)
    assert panel.columns.tolist() == (
        STRUCTURAL_OUTPUT_COLUMNS + PRIMARY_OUTPUT_COLUMNS
    )


def test_sector_and_industry_follow_ticker_and_company(quarterly_payload):
    panel = _exported_panel(quarterly_payload)
    assert panel.columns.tolist()[:4] == [
        "symbol",
        "company",
        "sector",
        "industry",
    ]
    assert panel["sector"].eq("Technology").all()
    assert panel["industry"].eq("Software").all()


def test_sources_are_immediately_before_related_calculations(quarterly_payload):
    columns = _exported_panel(quarterly_payload).columns.tolist()
    assert columns[columns.index("tax_provision"):columns.index("calc_nopat_quarterly") + 1] == [
        "tax_provision",
        "calc_tax_expense_quarterly",
        "pretax_income",
        "calc_raw_tax_rate_quarterly",
        "ebit",
        "calc_nopat_quarterly",
    ]
    assert columns[columns.index("short_term_debt_and_capital_lease"):columns.index("valuations__ratios__debt_to_equity") + 1] == [
        "short_term_debt_and_capital_lease",
        "long_term_debt_and_capital_lease",
        "calc_debt_value_quarterly",
        "total_assets",
        "total_liabilities",
        "equity",
        "total_stockholders_equity",
        "calc_debt_to_equity_quarterly",
        "valuations__ratios__debt_to_equity",
    ]
    assert columns[columns.index("market_cap"):columns.index("valuations__valuation_ratios__enterprise_value_to_fcf") + 1] == [
        "market_cap",
        "calc_debt_value_quarterly",
        "cash_and_cash_equivalents",
        "short_term_investments",
        "calc_enterprise_value_quarterly",
        "free_cash_flow",
        "calc_free_cash_flow_ttm",
        "calc_ev_to_fcf_quarterly",
        "valuations__valuation_ratios__enterprise_value_to_fcf",
    ]


def test_only_shared_calculation_sources_are_duplicated(quarterly_payload):
    panel = _exported_panel(quarterly_payload)
    assert not EXCLUDED_OUTPUT_COLUMNS & set(panel.columns)
    counts = panel.columns.value_counts()
    assert counts[counts > 1].to_dict() == {
        # Debt feeds three calculations: the EV bridge, debt-to-equity, and
        # the WACC capital structure. Market cap feeds EV and WACC.
        "calc_debt_value_quarterly": 3,
        "cash_and_cash_equivalents": 2,
        "short_term_investments": 2,
        "market_cap": 2,
    }


def test_all_financial_values_are_float(quarterly_payload):
    panel = _exported_panel(quarterly_payload)
    for position, column in enumerate(panel.columns):
        if column in STRUCTURAL_OUTPUT_COLUMNS:
            continue
        if column in NON_NUMERIC_OUTPUT_COLUMNS:
            continue
        assert is_float_dtype(panel.iloc[:, position]), column


def test_classification_columns_survive_the_export_as_text(quarterly_payload):
    """The float coercion must skip the label columns.

    Without the exemption every classification would be coerced to NaN and the
    workbook would look complete while carrying no decomposition at all.
    """
    panel = _exported_panel(quarterly_payload)
    labels = panel["calc_roic_business_classification"]
    assert labels.notna().all()
    assert labels.map(type).eq(str).all()
    assert panel["calc_roic_explanation"].str.len().gt(0).all()


def test_excel_number_formats_match_the_column_kind(
    quarterly_payload,
    tmp_path,
):
    panel = _exported_panel(quarterly_payload)
    path = tmp_path / "format.xlsx"
    write_excel(path, panel)
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook["Data"]
    for position, column in enumerate(panel.columns, start=1):
        expected = number_format_for(column) or "General"
        assert sheet.cell(row=2, column=position).number_format == expected, column


def test_every_column_keeps_full_precision_in_csv(quarterly_payload, tmp_path):
    """Rounding on write would change the number, not just its display.

    A ROIC contribution is O(1e-3), so "%.2f" used to export it as 0.00 and a
    per-column exemption list existed to rescue it. Without any float_format
    the exemption is unnecessary and no column loses digits.
    """
    panel = _exported_panel(quarterly_payload).copy()
    panel["calc_roic_ebit_contribution"] = 0.00312345
    panel["market_cap"] = 313.3299865722656
    path = tmp_path / "precision.csv"
    write_csv(path, panel)
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    header, first = rows[0], rows[1]
    assert first[header.index("calc_roic_ebit_contribution")] == "0.00312345"
    assert first[header.index("market_cap")] == "313.3299865722656"


def test_dummy_columns_move_to_their_own_excel_sheet(quarterly_payload, tmp_path):
    panel = _exported_panel(quarterly_payload)
    data, dummies = split_decomposition_dummies(panel)

    assert not set(DECOMPOSITION_DUMMY_COLUMNS) & set(data.columns)
    assert set(DECOMPOSITION_DUMMY_COLUMNS) <= set(dummies.columns)
    assert dummies.columns.tolist()[:2] == ["symbol", "period_key"]

    path = tmp_path / "split.xlsx"
    write_excel(path, data, decomposition=dummies)
    workbook = load_workbook(path)
    assert "Decomposition" in workbook.sheetnames
    assert workbook["Data"].max_column == data.shape[1]


def test_csv_and_excel_data_sheet_are_the_same_table(quarterly_payload, tmp_path):
    """הדרישה במלואה: אותן עמודות, אותו סדר ואותם ערכים בשני הקבצים.

    שניהם נכתבים מ-``data_sheet`` — הפאנל ללא עמודות הדמה — ושניהם נקראים
    בחזרה לאותה טבלה בדיוק. עיגול בכתיבה או עמודה שנוספה רק לאחד מהם ייפול
    כאן.
    """
    data_sheet, dummies = split_decomposition_dummies(
        _exported_panel(quarterly_payload)
    )
    csv_path, excel_path = tmp_path / "panel.csv", tmp_path / "panel.xlsx"
    write_csv(csv_path, data_sheet)
    write_excel(excel_path, data_sheet, decomposition=dummies)

    # ⚠️ float_precision="round_trip" חיוני: הפרסר המהיר שהוא ברירת המחדל של
    # read_csv אינו מעוגל נכון ומחזיר כ-13 ספרות מובהקות בלבד. בלעדיו הבדיקה
    # הייתה מודדת את הפרסר במקום את הקובץ, וסוטה ב-8e-13 גם כשהקובץ מדויק.
    from_csv = pd.read_csv(
        csv_path, encoding="utf-8-sig", float_precision="round_trip"
    )
    from_excel = pd.read_excel(excel_path, sheet_name="Data")

    assert from_csv.columns.tolist() == from_excel.columns.tolist()
    # הסבילות נבחרה ולא הושארה כברירת מחדל, משתי סיבות מנוגדות:
    #
    #   ברירת המחדל (rtol=1e-5) רופפת מדי — עיגול לשתי ספרות של
    #   313.3299865722656 סוטה ב-4e-8 בלבד והיה עובר אותה בשקט, כלומר
    #   הבדיקה לא הייתה מסוגלת להיכשל על מה שהיא נועדה לתפוס.
    #
    #   check_exact=True הדוק מדי — פורמט xlsx שומר 16 ספרות מובהקות ולא 17,
    #   ולכן 1.8828235607599488 חוזר ממנו כ-1.882823560759949. ה-CSV הוא
    #   הצד המדויק כאן; ההפרש הוא ULP בודד.
    pd.testing.assert_frame_equal(
        from_csv, from_excel, check_dtype=False, rtol=1e-15, atol=0
    )


def test_dummy_columns_are_absent_from_both_data_outputs(quarterly_payload):
    """עמודות הדמה יושבות בגיליון Decomposition ובפלט parquet בלבד."""
    data_sheet, dummies = split_decomposition_dummies(
        _exported_panel(quarterly_payload)
    )
    assert not set(DECOMPOSITION_DUMMY_COLUMNS) & set(data_sheet.columns)
    assert set(DECOMPOSITION_DUMMY_COLUMNS) <= set(dummies.columns)
